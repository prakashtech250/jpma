from openpyxl import Workbook, load_workbook
import scrapy
import requests
import re
import csv
import os

class SpiderSpider(scrapy.Spider):
    name = 'spider'
    allowed_domains = ['jpma.org.pk']
    start_urls = ['http://www.jpma.org.pk/IssuesList']
    base_url = 'https://www.jpma.org.pk'
    article_count = 1
    done = []
    title = ['Sr.No', 'Year', 'Month','Journal' ,'Volume', 'Issue', 'Editorial', 'Editorial Author', 'Article','Article Author','Article Author Af', 'Letter', 'Letter Author', 'Letter Author Af','Student', 'Student_Author','Student Author Af']
    if os.path.exists('done.txt'):
        with open('done.txt','r') as f:
            done=[i.strip() for i in f]
    else:
        pass
    count = {}
    if not os.path.exists('output'):
        os.mkdir('output')

    def parse(self, response):
        link_divs = response.css('#content a')
        for link_div in link_divs:
            link = link_div.css('::attr(href)').get()
            if 'past-issue' not in link:
                continue
            title = link_div.css('::text').get().strip()
            print(title)
            year = title.split()[-1]
            self.count[year] = 0
            filename = f'output/{year}.xlsx'
            if os.path.exists(filename):
                self.wb = load_workbook(filename)
                self.ws = self.wb.active
            else:
                self.wb = Workbook()
                self.ws = self.wb.active
                self.ws.append(self.title)
            url = response.urljoin(link)
            yield scrapy.Request(url, callback=self.parse_details, meta={
                'filename': filename,
            })

    def get_details(self, url):
        while True:
            response = requests.get(url)
            if response.status_code == 200:
                sel = scrapy.Selector(text=response.text)
                texts = sel.css('.entry-title+ .entry-content p::text').extract()
                article = sel.css('.entry-title b::text').extract_first()
                print(f'{self.article_count}: {article}')
                self.article_count += 1
                authors = []
                authors_af = []
                for text in texts:
                    regex = r"(?=\()(.*)"
                    try:
                        rex = re.search(regex, text).group(0)
                        author = text.replace(rex,'').strip()
                    except:
                        author = text.strip()
                    Af_regex = r"(?<=\()(.*)(?=\))"
                    try:
                        Af = re.search(Af_regex, text).group(0)
                    except:
                        Af = None
                    if Af:
                        authors_af.append(Af)
                    if author:
                        authors.append(author)
                if len(authors) > 0:
                    auth = ';'.join(authors)
                else:
                    auth = None
                if len(authors_af) > 0:
                    au_af = ';'.join(authors_af)
                else:
                    au_af = None
                return (article, auth , au_af)
            else:
                print(f'{self.article_count}: Trying again.... [{response.status_code}]')
        

    def parse_details(self, response):
        filename = response.meta['filename']
        if not os.path.exists(filename):
            with open(filename, 'w') as f:
                writer = csv.writer(f)
                writer.writerow(self.title)

        title = response.css('.title-style::text').get().strip()
        year = title.split(',')[0].split()[1]
        month = title.split(',')[0].split()[0]
        volume = title.split(',')[1].split()[1]
        issue = title.split(',')[2].split()[1]
        print(f'{self.count[year]}: {title}')
        print('Past Issue: ', response.url)
        editorial = response.css('#content .clearfix .clearfix .col-md-12 b::text').get().strip()
        editorial_authors = response.css('.author-italic p::text').get().strip()
        banner = response.css('.origionalbg')
        orignal_articles_links = banner.css('#carousel-example-generic p a::attr(href)').extract()
        scrolls = response.css('.col-md-4')
        research_articles_links = []
        issue_articles_links = []
        review_articles_links = []
        case_articles_links = []
        letter_articles_links = []
        students_articles_links = []
        articles = []
        letters = []
        students = []
        for scroll in scrolls:
            title = scroll.css('.fancy-title h3::text').get()
            if title == 'RESEARCH ARTICLES':
                research_articles_links = scroll.css('.articleScroll a::attr(href)').extract()
                research_articles_links = list(set(research_articles_links))
                research_articles_links = [x for x in research_articles_links if 'Download' not in x]
            elif title == 'IN THIS ISSUE':
                issue_articles_links = scroll.css('.articleScroll a::attr(href)').extract()
                issue_articles_links = list(set(issue_articles_links))
                issue_articles_links = [x for x in issue_articles_links if 'Download' not in x]
            elif title == 'REVIEW ARTICLES':
                review_articles_links = scroll.css('.articleScroll a::attr(href)').extract()
                review_articles_links = list(set(review_articles_links))
                review_articles_links = [x for x in review_articles_links if 'Download' not in x]
            elif title == 'CASE REPORTS':
                case_articles_links = scroll.css('.articleScroll a::attr(href)').extract()
                case_articles_links = list(set(case_articles_links))
                case_articles_links = [x for x in case_articles_links if 'Download' not in x]
            elif title == 'LETTER TO THE EDITOR':
                letter_articles_links = scroll.css('.articleScroll a::attr(href)').extract()
                letter_articles_links = list(set(letter_articles_links))
                letter_articles_links = [x for x in letter_articles_links if 'Download' not in x]
            elif title == 'STUDENTS\' CORNER':
                students_articles_links = scroll.css('.articleScroll a::attr(href)').extract()
                students_articles_links = list(set(students_articles_links))
                students_articles_links = [x for x in students_articles_links if 'Download' not in x]
        article_links = orignal_articles_links + research_articles_links + issue_articles_links + review_articles_links + case_articles_links
        article_links = [self.base_url + x for x in article_links]
        for article_link in article_links:
            print(article_link)
            article = self.get_details(article_link)
            articles.append(article)
        letter_articles_links = [self.base_url + x for x in letter_articles_links]
        for letter_article_link in letter_articles_links:
            letter = self.get_details(letter_article_link)
            letters.append(letter)
        students_articles_links = [self.base_url + x for x in students_articles_links]
        for students_article_link in students_articles_links:
            student = self.get_details(students_article_link)
            students.append(student)
        i = 0
        for article in articles:
            article_text = article[0]
            article_author = article[1]
            article_af = article[2]
            try:
                letter_text = letters[i][0]
                letter_author = letters[i][1]
                letter_author_af = letters[i][2]
            except:
                letter_text = None
                letter_author = None
                letter_author_af = None
            try:
                student_text = students[i][0]
                student_author = students[i][1]
                student_author_af = students[i][2]
            except:
                student_text = None
                student_author = None
                student_author_af = None
            if i == 0:
                pass
            else:
                editorial = None
                editorial_authors = None         
            i += 1
            # data = [self.count[year], year, month, volume, issue, editorial, editorial_authors, article_text, article_author, letter_text, letter_author, student_text, student_author]
            self.count[year] += 1

            scraped_info = {
                'Sr.no': self.count[year],
                'Year': year,
                'Month': month,
                'Journal': 'JPMA',
                'Volume': volume,
                'Issue': issue,
                'Editorial': editorial,
                'Editorial Author': editorial_authors,
                'Article': article_text,
                'Article Author': article_author,
                'Article Author Af': article_af,
                'Letter': letter_text,
                'Letter Author': letter_author,
                'Letter Author Af': letter_author_af,
                'Student': student_text,
                'Student Author': student_author,
                'Student Author Af': student_author_af,
            }
            yield scraped_info

            data = list(scraped_info.values())
            with open(filename, 'a') as f:
                writer = csv.writer(f)
                writer.writerow(data)
            self.ws.append(data)
            self.wb.save(filename)
        
